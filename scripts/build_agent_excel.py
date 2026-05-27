"""Build the consolidated AI Agent funding xlsx (past 6 months: 2025-11 ~ 2026-05)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROWS = [
    # Layer, Company, Period, Round, Amount, Valuation, Lead Investor, Direction, Source
    ("模型", "OpenAI", "01-22~02-27", "growth", "$110B", "$840B", "Amazon / Nvidia / SoftBank", "前沿通用大模型（ChatGPT）", ""),
    ("模型", "Anthropic", "01-08~05-15", "VC", "$50B", "$850-900B", "—", "Claude", ""),
    ("模型", "xAI", "01-08", "Series E", "$20B", "—", "Valor Equity 等", "Grok（已并入 SpaceX）", ""),
    ("模型", "Moonshot AI 🇨🇳", "05-08", "Growth", "$2B", "$20B", "Meituan", "Kimi 长文本大模型", ""),
    ("模型", "Reflection AI", "03-03", "Late stage", "$2B", "$20B", "—", "自主编码/通用 agent 研究", ""),
    ("模型", "Ineffable Intelligence", "04-28", "seed", "$1.1B", "$5.1B", "Sequoia / Lightspeed", "RL 驱动前沿模型", ""),
    ("模型", "AMI Labs", "03-10", "venture", "$1.03B", "$3.5B", "—", "世界模型", ""),
    ("模型", "Recursive Superintelligence", "05-14", "Growth", "$650M", "$4.65B", "GV / Greycroft", "通用超级智能", ""),
    ("模型", "ElevenLabs", "02-05", "Series D", "$500M", "$11B", "Sequoia", "语音合成", ""),
    ("模型", "humans&", "01-20", "seed", "$480M", "$4.48B", "SV Angel; Georges Harik", "通用 AI 实验室", ""),
    ("模型", "Sarvam 🇮🇳", "04-03", "growth", "$350M", "~$1.5B", "Bessemer", "印度主权大模型", ""),
    ("模型", "Fal", "03-20", "venture", "$325M", "~$8B", "—", "生成式媒体推理云", ""),
    ("模型", "Runway", "02-10", "venture", "$315M", "$5.3B", "General Atlantic", "AI 视频生成", ""),
    ("模型", "PixVerse 🇨🇳", "03-12", "Series C", "$300M", "$1B", "CDH Investments", "AI 视频（爱诗科技）", ""),
    ("模型", "ShengShu 🇨🇳", "04-10", "growth", "$293M", "—", "Alibaba", "Vidu 视频大模型", ""),
    ("模型", "Wispr", "05-13", "Growth", "$260M", "$2B", "Menlo", "语音输入（Flow）", ""),
    ("模型", "Fundamental", "02-06", "Series A", "$225M", "$1.4B post", "Oak HC/FT", "企业基础模型", ""),
    ("模型", "Axiom Math", "03-13", "Series A", "$200M", "$1.6B", "Menlo", "数学超智能", ""),
    ("模型", "Goodfire", "02-06", "Series B", "$150M", "$1.25B", "B Capital", "模型可解释性", ""),
    ("模型", "LMArena", "01-07", "growth", "$150M", "$1.7B", "Felicis / UC Investments", "模型评测竞技场", ""),
    ("模型", "Deepgram", "01-13", "Series C", "$130M", "$1.3B", "Alkeon 等", "语音识别 API", ""),
    ("模型", "Harmonic", "01-15~01-26", "Series C", "$120M", "$1.45B", "NVentures / Ribbit / Sequoia", "数学推理", ""),
    ("模型", "Higgsfield", "01-16", "Series A ext", "$80M", "$1.3B", "Accel / GFT / Menlo", "AI 视频（创作者）", ""),
    ("模型", "Standard Intelligence", "05-01", "VC", "$75M", "$500M", "Sequoia / Spark", "视频训练通用模型", ""),
    ("模型", "Mirage", "03-25", "growth", "$75M", "—", "General Catalyst", "AI 视频编辑", ""),
    ("模型", "webAI", "01-13", "Series A ext", "$75M", "$2.5B pre", "Time Ventures", "端侧/主权 AI", ""),
    ("模型", "Deccan AI", "03-26", "Series A", "$25M", "—", "A91 Partners", "post-training 数据", ""),
    ("模型", "Featherless", "04-30", "Series A", "$20M", "—", "AMD / Airbus Ventures", "开源模型托管", ""),
    ("模型", "Augur", "03-09", "venture", "$15M", "—", "Plural", "主权 AI", ""),
    ("模型", "Autoscience", "03-19", "seed", "$14M", "—", "General Catalyst", "AI 科研自动化", ""),
    ("模型", "Phylo", "02-04", "seed", "$13.5M", "—", "Menlo / a16z", "AI 研究实验室", ""),
    ("模型", "AgileRL", "01-08", "early", "$7.5M", "—", "Fusion Fund", "RL 训练框架", ""),
    ("模型", "Berget AI", "02-03", "venture", "$2.3M", "—", "Luminar Ventures", "欧洲主权 AI 云", ""),
    ("模型", "RadixArk (SGLang)", "01-22", "venture", "—", "$400M", "Accel", "开源推理引擎", ""),

    ("Agent 基础设施", "Databricks", "02-10", "Series L", "$5B", "$134B", "Insight / Fidelity / JPMorgan", "数据+AI 平台", ""),
    ("Agent 基础设施", "Replit", "03-12", "Series D", "$400M", "—", "Georgian", "AI 云端开发环境", ""),
    ("Agent 基础设施", "Exa Labs", "05-21", "growth", "$250M", "$2.2B", "a16z", "AI 原生搜索 API", ""),
    ("Agent 基础设施", "Inferact", "Q1", "Seed", "$150M", "$800M", "a16z / Lightspeed", "agent inference 基础设施", "🆕 Web"),
    ("Agent 基础设施", "Granola", "03-26", "venture", "$125M", "$1.5B", "Index Ventures", "AI 会议记录", ""),
    ("Agent 基础设施", "Port", "Q1", "venture", "$100M", "—", "General Atlantic / Accel", "Agentic engineering 平台", "🆕 Web"),
    ("Agent 基础设施", "Coder", "04-01", "Series C", "$90M", "—", "KKR", "云端开发环境（自托管）", ""),
    ("Agent 基础设施", "Braintrust", "02-17", "Series B", "$80M", "$800M post", "Iconiq", "LLM 评测与可观测", ""),
    ("Agent 基础设施", "Qodo", "03-31", "Series B", "$70M", "—", "Qumra Capital", "AI 代码质量/测试", ""),
    ("Agent 基础设施", "Sycamore Labs", "03-30", "seed", "$65M", "—", "Coatue / Lightspeed", "AI agent 操作系统", ""),
    ("Agent 基础设施", "Gumloop", "03-13", "Series B", "$50M", "—", "Benchmark", "无代码 agent 搭建", ""),
    ("Agent 基础设施", "Vapi", "05-12", "venture", "未披露", "$500M", "—", "Voice Agent 基础设施", "🆕 Web"),
    ("Agent 基础设施", "Nimble", "Q1", "venture", "$47M", "—", "Norwest / Databricks Ventures", "agent 用 Web 搜索", "🆕 Web"),
    ("Agent 基础设施", "Deeptune", "03-19", "Series A", "$43M", "—", "a16z", "agent 训练/微调", ""),
    ("Agent 基础设施", "Onyx Security", "03-13", "Series A", "$35M", "—", "Conviction", "agent 安全", ""),
    ("Agent 基础设施", "Parasail", "04-15", "Series A", "$32M", "—", "Touring / Kindred", "agent/模型部署算力", ""),
    ("Agent 基础设施", "Linq", "02-03", "Series A", "$20M", "—", "TQ / Mucker", "企业 agent 基础设施", ""),
    ("Agent 基础设施", "Interloop", "03-24", "seed", "$17M", "—", "DN Capital", "agent 记忆层", ""),
    ("Agent 基础设施", "Moonbounce", "04-06", "venture", "$12M", "—", "Amplify / StepStone", "agent 控制引擎", ""),
    ("Agent 基础设施", "Whirl AI", "03-31", "seed", "$8.9M", "—", "Iconiq", "企业知识接入", ""),
    ("Agent 基础设施", "Nava", "04-15", "seed", "$8.3M", "—", "Polychain / Archetype", "agent 经济护栏/合规", ""),
    ("Agent 基础设施", "Manufact", "02-12", "seed", "$6.3M", "—", "Peak XV", "MCP agent 基础设施", ""),
    ("Agent 基础设施", "Neuramint", "01-13", "seed", "$5M", "—", "Maelstrom 等", "agent 平台", ""),
    ("Agent 基础设施", "Certiv", "03-17", "pre-seed", "$4.2M", "—", "Aviso / Founders' Co-op", "agent 安全/认证", ""),
    ("Agent 基础设施", "Modern Relay", "04-15", "VC", "$3M", "—", "Point Nine", "企业 agent 基础设施", ""),
    ("Agent 基础设施", "Laminar", "03-18", "seed", "$3M", "—", "Atlantic.vc", "agent 调试/追踪", ""),
    ("Agent 基础设施", "Ralio", "04-15", "VC", "$2.5M", "—", "Sure Valley Ventures", "AI agent 支付/结算", ""),

    ("Agent 应用", "Cursor (Anysphere)", "04", "growth（谈判中）", "$2B", "$50B", "a16z / Thrive Capital", "AI 代码编辑器", "🆕 Web"),
    ("Agent 应用", "Sierra", "05-05", "Growth", "$950M", "$15.8B", "Tiger Global / GV", "企业客服 agent（Bret Taylor）", ""),
    ("Agent 应用", "Legora", "03-10~04-30", "Series D", "$550M", "$5.55B", "Accel", "法律工作流 agent", ""),
    ("Agent 应用", "Cognition", "2025-09 ⏪", "Series C", "$400M", "$10.2B", "Founders Fund", "自主软件工程师（Devin）", "🆕 Web"),
    ("Agent 应用", "Mercor", "Q1", "venture", "$350M", "—", "Felicis / Benchmark", "AI 招聘/匹配", "🆕 Web"),
    ("Agent 应用", "Parloa", "01-15~01-26", "Series D", "$350M", "$3B", "General Catalyst", "客服语音 agent", ""),
    ("Agent 应用", "OpenEvidence", "Q1", "Series D", "$250M", "$12B", "Thrive Global / DST Global", "临床决策医疗 AI", "🆕 Web"),
    ("Agent 应用", "Decagon", "01-29", "Series D", "$250M", "$4.5B", "Coatue / Index", "客服 concierge", ""),
    ("Agent 应用", "Harness", "2025-Q4 ⏪", "Series E", "$240M", "$5.5B", "Goldman Sachs", "DevOps + AI agent", "🆕 Web"),
    ("Agent 应用", "Blitzy", "05-06", "Growth", "$200M", "$1.4B", "Northzone", "自主软件开发", ""),
    ("Agent 应用", "Harvey", "03-26", "venture", "$200M", "$11B", "GIC / Sequoia", "法律 AI 助手", ""),
    ("Agent 应用", "Rogo", "01-28~04-30", "Series D", "$160M", "—", "Kleiner Perkins", "投行/PE 金融分析", ""),
    ("Agent 应用", "Hightouch", "04-30", "Series D", "$150M", "$2.75B", "Goldman Sachs Alt / Bain Cap", "Composable CDP 营销", ""),
    ("Agent 应用", "Wonderful", "03-13", "Series B", "$150M", "$2B", "Insight Partners", "多语种客服", ""),
    ("Agent 应用", "Glean", "Q1", "Series F", "$150M", "$7.2B", "Wellington Management", "企业搜索 + agent", "🆕 Web"),
    ("Agent 应用", "EvenUp", "Q1", "venture", "$150M", "—", "Bessemer / REV", "人身伤害法 AI", "🆕 Web"),
    ("Agent 应用", "Factory", "04", "venture", "$150M", "$1.5B", "Khosla Ventures", "企业级 AI Coding", "🆕 Web"),
    ("Agent 应用", "7AI", "2025-12-04 ⏪", "Series A", "$130M", "—", "Index Ventures（Greylock/Spark/CRV）", "网络安全 agent", "🆕 Web"),
    ("Agent 应用", "Hippocratic AI", "2025-11 ⏪", "Series C", "$126M", "$3.5B", "Avenir Growth / CapitalG", "医疗对话 agent", "🆕 Web"),
    ("Agent 应用", "Resolve AI", "02-05~04-16", "venture", "$125M", "$1B", "Lightspeed", "SRE/运维 agent", ""),
    ("Agent 应用", "Kai", "03-11", "venture", "$125M", "—", "Evolution Equity Partners", "网络安全 agent", ""),
    ("Agent 应用", "Avoca", "04", "Series B", "$125M+", "$1B", "Meritech / General Catalyst", "Home Services agent", "🆕 Web"),
    ("Agent 应用", "Oasis Security", "Q1", "venture", "$120M", "—", "Craft / Cyberstarts / Sequoia / Accel", "AI agent 身份安全", "🆕 Web"),
    ("Agent 应用", "Netomi", "04-30", "VC", "$110M", "—", "Accenture Ventures", "客服 agent", ""),
    ("Agent 应用", "Isara", "03-26", "venture", "$94M", "$650M", "Amity Ventures", "企业 agents", ""),
    ("Agent 应用", "PolyAI", "2025-Q4 ⏪", "Series D", "$86M", "—", "Georgian / Hedosophia / Khosla", "语音客服", "🆕 Web"),
    ("Agent 应用", "Fieldguide", "02-02", "Series C", "$75M", "$700M", "Goldman Sachs Alt", "审计/咨询自动化", ""),
    ("Agent 应用", "Articul8", "Q1", "venture", "$70M", "—", "Adara Ventures / NXC", "企业 AI 平台", "🆕 Web"),
    ("Agent 应用", "JuliaHub", "04-30", "Series B", "$65M", "—", "Dorilton Capital", "科学计算/仿真", ""),
    ("Agent 应用", "Crosby", "03-31", "Series B", "$60M", "—", "Lux / Index", "法律合同审查", ""),
    ("Agent 应用", "Surf AI", "03-17", "venture", "$57M", "—", "Accel", "SOC 安全运营", ""),
    ("Agent 应用", "Ivo", "01-20", "Series B", "$55M", "—", "Blackbird", "合同审查（法律）", ""),
    ("Agent 应用", "Steno", "03-26", "Series C", "$49M", "—", "Savano Capital", "庭审记录", ""),
    ("Agent 应用", "Patlytics", "04-08", "Series B", "$40M", "—", "SignalFire", "专利分析", ""),
    ("Agent 应用", "RunSybil", "03", "venture", "$40M", "—", "Khosla Ventures", "AI 渗透测试", "🆕 Web"),
    ("Agent 应用", "Solve Intelligence", "Q1", "venture", "$40M", "—", "Visionaries Club / 20VC", "专利工作流", "🆕 Web"),
    ("Agent 应用", "Glimpse", "03-25", "Series A", "$35M", "—", "a16z", "零售/CPG 消费洞察", ""),
    ("Agent 应用", "Monaco", "02-12", "seed/Series A", "$35M", "—", "Founders Fund", "销售", ""),
    ("Agent 应用", "Sett", "03-30", "Series B", "$30M", "—", "Greenfield Partners", "游戏内容/运营", ""),
    ("Agent 应用", "Notch", "03-25", "Series A", "$30M", "—", "Headline", "客户体验", ""),
    ("Agent 应用", "Dify 🇨🇳", "03-10", "venture", "$30M", "$180M", "HSG", "开源 LLM 应用/agent 工作流", ""),
    ("Agent 应用", "Didero", "02-13", "Series A", "$30M", "—", "Chemistry / Headline", "采购 agent", ""),
    ("Agent 应用", "Catena Labs", "05-20", "Series A", "$30M", "—", "Acrew Capital", "AI-native Bank", "🆕 Web"),
    ("Agent 应用", "1mind", "Q1", "venture", "$30M", "—", "Battery / Primary", "AI Sales 平台", "🆕 Web"),
    ("Agent 应用", "Uptiq", "02-12", "Series B", "$25M", "—", "Curql", "金融机构信贷 agent", ""),
    ("Agent 应用", "Newo", "02-11", "Series A", "$25M", "—", "Ratmir Timashev", "语音 agent", ""),
    ("Agent 应用", "DeepIP", "Q1", "venture", "$25M", "—", "Korelya / Serena", "专利起草", "🆕 Web"),
    ("Agent 应用", "Resistant AI", "2025-11 ⏪", "Series B", "$25M", "—", "DTCP", "金融犯罪检测", "🆕 Web"),
    ("Agent 应用", "Parallel", "03-19", "Series A", "$20M", "—", "Index Ventures", "医疗 agent", ""),
    ("Agent 应用", "Day AI", "02-03", "Series A", "$20M", "—", "Sequoia", "AI CRM（邮件/日历）", ""),
    ("Agent 应用", "Flip", "01-23", "Series A", "$20M", "—", "Next Coast / Ridge", "客服", ""),
    ("Agent 应用", "Meridian", "02-12", "seed", "$17M", "$100M post", "a16z / General Partnership", "Agentic 电子表格", ""),
    ("Agent 应用", "Qurrent", "03-12", "Series A", "$15M", "—", "Cervin Ventures", "数字员工", ""),
    ("Agent 应用", "EnFi", "02-09", "venture", "$15M", "—", "Fintop", "银行信贷分析", ""),
    ("Agent 应用", "Sapiom", "02-06", "seed", "$15M", "—", "Accel", "采购 agents", ""),
    ("Agent 应用", "Synthpop", "02-03", "Series A", "$15M", "—", "Ansa", "医疗行政/理赔", ""),
    ("Agent 应用", "Simple AI", "02-11", "seed", "$14M", "—", "First Harmonic", "语音 agent（代打电话）", ""),
    ("Agent 应用", "Level3AI", "01-22", "seed", "$13M", "—", "Lightspeed", "客服质检", ""),
    ("Agent 应用", "Definity", "04-30", "Series A", "$12M", "—", "GreatPoint Ventures", "agentic 数据工程", ""),
    ("Agent 应用", "Mega", "03-09", "Series A", "$11.5M", "—", "Goodwater Capital", "营销", ""),
    ("Agent 应用", "RISA Labs", "01-13", "Series A", "$11.1M", "—", "Cencora / Optum Ventures", "肿瘤科", ""),
    ("Agent 应用", "Amigo AI", "03-10", "Series A", "$11M", "—", "Madrona", "患者端", ""),
    ("Agent 应用", "General Analysis", "04-29", "seed", "$10M", "—", "Altos Ventures", "agent 安全/红队", ""),
    ("Agent 应用", "Capsule", "04-16", "seed", "$7M", "—", "Lama / Forgepoint Intl", "agentic 安全", ""),
    ("Agent 应用", "Obin AI", "03-18", "seed", "$7M", "—", "Motive Partners", "金融 agents", ""),
    ("Agent 应用", "Miravoice", "04-03", "seed", "$6.3M", "—", "Unusual Ventures", "语音 agent", ""),
    ("Agent 应用", "Airrived", "02-04", "seed", "$6.1M", "—", "Cannage Capital", "企业 agents", ""),
    ("Agent 应用", "Toyo", "02-17", "seed", "$4.3M", "—", "Frontline Ventures", "AI agents", ""),
    ("Agent 应用", "Handhold", "04-09", "seed", "$3.3M", "—", "Entourage Capital", "销售", ""),
    ("Agent 应用", "Riplo", "04-01", "pre-seed", "$3M", "—", "Cherry Ventures", "agentic 咨询", ""),
    ("Agent 应用", "GrowthLoop", "01-14", "growth", "—", "—", "TJC", "agentic 营销/数据", ""),
]

HEADERS = ["层级", "公司", "时间", "轮次", "金额", "估值", "领投", "方向", "来源"]

wb = Workbook()
ws = wb.active
ws.title = "AI Agent 融资 2025-11~2026-05"

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
layer_fills = {
    "模型": PatternFill("solid", fgColor="FFF2CC"),
    "Agent 基础设施": PatternFill("solid", fgColor="E2EFDA"),
    "Agent 应用": PatternFill("solid", fgColor="DDEBF7"),
}
web_font = Font(color="C00000", bold=True)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADERS)
for c, _ in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for row in ROWS:
    ws.append(row)
    r = ws.max_row
    fill = layer_fills.get(row[0])
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill
    if row[8].startswith("🆕"):
        ws.cell(row=r, column=9).font = web_font

widths = [16, 26, 16, 18, 12, 14, 38, 38, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

notes_ws = wb.create_sheet("说明")
notes = [
    ["合并表说明"],
    [],
    ["统计窗口", "2025-11 ~ 2026-05（过去约 6 个月）"],
    ["数据条数", f"{len(ROWS)} 家（模型 34 + Agent 基础设施 27 + Agent 应用 71）"],
    [],
    ["来源标记"],
    ["空白", "本地 data/2026-Jan-May.json（Axios Pro Rata 主源）"],
    ["🆕 Web", "Web 补充（24 家）"],
    ["⏪", "早于本地 Jan-May 窗口，但仍在过去半年内；Cognition 2025-09 按用户要求保留"],
    [],
    ["未含"],
    ["", "芯片/算力层（21 家）"],
    ["", "机器人 / 物理 AI（Shield AI / Halter / Rhoda 等）"],
    ["", "健康/生物科技降权赛道（按 preferences.md）"],
    [],
    ["待核对冲突"],
    ["Parallel", "本地'医疗 agent' vs Web 'Web infra' — 高度可能是同名两家公司"],
    ["Rogo", "本地 $160M Series D（Kleiner Perkins）vs Web $75M（Sequoia + Henry Kravis）— 不同口径"],
]
for r in notes:
    notes_ws.append(r)
notes_ws.column_dimensions["A"].width = 20
notes_ws.column_dimensions["B"].width = 80
notes_ws["A1"].font = Font(bold=True, size=14)

out = "/home/user/Startup-Tracking/data/2025-11_2026-05_AI-agent-funding.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {len(ROWS)}")
